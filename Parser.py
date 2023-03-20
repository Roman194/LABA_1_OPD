from bs4 import BeautifulSoup # импортируем библиотеку BeautifulSoup
import requests # импортируем библиотеку requests
from openpyxl import load_workbook #импорт библиотеки для работы с таблицей

def parse():
    my_list='Vacansii_list.xlsx'
    lwb=load_workbook(my_list)      #список листов файла
    ws=lwb['данные']    #работа с конкретным листом файла
    url = 'https://omsk.rabota.ru/?query=Python&sort=relevance'     #передаем необходимы URL адрес

    try:
        page = requests.get(url)    #отправляем запрос методом Get на данный адрес и получаем ответ в переменную
    except ConnectionError:
        print('Не удалось спарсить. Похоже есть защита у сайта')
        return -1

    print(page.status_code) # смотрим ответ

    block = []
    filtered_block_name=[]
    filtered_block_saly=[]
    filtered_block_desc=[]
    filtered_block_comp=[]
    filtered_block_woty=[]
    list_of_hat_of_table=['Должность','Зарплата','Описание','Работодатель','Тип работы']

    for i in range(1,6):    #оформляем в... кхм шапку таблицы
        cell=ws.cell(2,i+1)
        cell.value=list_of_hat_of_table[i-1]

    soup = BeautifulSoup(page.text, "html.parser") # передаем страницу в bs4

    block = soup.findAll('div', class_='vacancy-preview-card__top' or 'r-serp-similar-title r-home-serp__item')         # находим  контейнера с нужными классами

    for data in block:      # проходим циклом по содержимому контейнера
        if data.find('h2'):     #class_='r-serp-similar-title r-home-serp__item'
            break       #это должно работать и не давать уходить в похожие вакансии
        else:
                # находим нужный тег <p> для разных типов данных
                name=data.find(class_='vacancy-preview-card__title_border')
                salary=data.find(class_='vacancy-preview-card__salary')
                description=data.find(class_='vacancy-preview-card__short-description')
                company=data.find(class_='vacancy-preview-card__company-name')
                work_type=data.find(class_='vacancy-preview-location__address-text')
                if (name and salary and description and company and work_type) is not None:
                    name_st=str((name.text).replace('\n',''))
                    name_st=name_st.replace('            ','')      #гениальный способ убрать некрасивые и лишние пробелы
                    name_st=name_st.replace('  ','')
                    if 'Водитель' in name_st:       #этот костыль ужасен но зато всё работает
                        break
                    filtered_block_name.append(name_st)     #записываем в переменную списка содержание тега

                    saly_st=str((salary.text).replace('\n',''))
                    saly_st = saly_st.replace('            ','')
                    saly_st = saly_st.replace('  ', '')
                    filtered_block_saly.append(saly_st)

                    desc_st=str((description.text).replace('\n',''))
                    desc_st = desc_st.replace('            ','')
                    desc_st = desc_st.replace('  ', '')
                    filtered_block_desc.append(desc_st)

                    comp_st=str((company.text).replace('\n',''))
                    comp_st= comp_st.replace('            ','')
                    comp_st= comp_st.replace('  ', '')
                    filtered_block_comp.append(comp_st)

                    work_st=str((work_type.text).replace('\n',''))
                    work_st=work_st.replace('            ','')
                    work_st=work_st.replace('  ', '')
                    filtered_block_woty.append(work_st)

    print(filtered_block_name,filtered_block_saly,filtered_block_desc,filtered_block_comp,filtered_block_woty)

    counter=3
    #для каждой переменной каждого списка записываем данные в собственную ячейку в таблице
    for elem_n, elem_s, elem_d, elem_c, elem_w in zip(filtered_block_name,filtered_block_saly,filtered_block_desc,filtered_block_comp,filtered_block_woty):
        cell=ws.cell(counter,2)
        elm=''.join(elem_n)     #преобразуем элл-ты списка ({0!r}) в строку (str)
        cell.value=elm

        cell=ws.cell(counter,3)
        elm=''.join(elem_s)
        cell.value=elm

        cell=ws.cell(counter,4)
        elm = ''.join(elem_d)
        cell.value=elm

        cell=ws.cell(counter,5)
        elm = ''.join(elem_c)
        cell.value=elm

        cell=ws.cell(counter,6)
        elm = ''.join(elem_w)
        cell.value = elm

        counter+=1

    lwb.save(my_list) #сохраняем и закрываем лист
    lwb.close()